"""
scripts/enrich_snopes_claims.py

Filter, deduplicate, and LLM-enrich a raw Snopes fact-check export into
the JSON schema ``SnopesDataset`` (``eiger.datasets.snopes``) expects.

Why this is a separate, offline script rather than logic inside
SnopesDataset.load()
----------------------------------------------------------------------
``BaseDataset.load()`` is expected to be fast and dependency-light (see
``JSONFixtureDataset``, ``SnopesDataset``'s own parent). Generating a
``context_query`` for thousands of claims via a live LLM is neither: it
can take a long time and requires a running Ollama server. Doing it here,
once, and caching the result to disk keeps ``SnopesDataset.load()`` fast,
deterministic, and offline — exactly like every other ``BaseDataset``
implementation in the project.

What this script does
-----------------------
1. Reads the raw Snopes export (e.g. ``Snopes.xlsx``). Column NAMES must
   include: ``claim_id``, ``claim``, ``url``, ``date_published``,
   ``original_verdict``, ``normalised_rating``; column order is not
   assumed, only names (read from the header row of the first sheet).
2. Keeps only rows where ``normalised_rating`` is exactly ``True``
   (verified-true claims — ``Claim.original_fact`` must be a true
   statement; EIGER generates its own falsehoods via the attack registry,
   it does not import externally-sourced false claims as ground truth).
3. Deduplicates by ``claim_id`` (first occurrence wins) — the raw export
   has been observed to contain a small number of duplicate claim_ids.
4. For each surviving claim, generates a natural-language question via a
   local Ollama LLM that the claim would directly answer.
5. Writes (or incrementally appends to) a JSON file in the schema
   ``SnopesDataset``/``JSONFixtureDataset`` expect: ``claim_id``,
   ``original_fact``, ``context_query``, ``source``, ``notes``,
   ``verified`` (always ``False`` on first write — Snopes' own "True"
   rating is not a substitute for this research team's own review; see
   ``scripts/README.md`` and ``eiger/datasets/snopes.py``).

Resumability
-------------
Enriching thousands of claims against a local LLM can take a long time
and may be interrupted. This script is idempotent: if the output file
already exists, already-enriched claim_ids are loaded and skipped, and
newly-enriched claims are appended rather than overwriting them. Progress
is also checkpointed to disk every ``--checkpoint-every`` claims (default
25), so an interruption loses at most that many claims' worth of work.

Usage
------
    # Requires: `pip install 'eiger[data-import]'` and a running Ollama
    # server with the target model pulled (`make up`, then
    # `docker exec eiger-ollama ollama pull llama3.1:8b`).

    # Pilot run on a small sample first (recommended before the full batch):
    python scripts/enrich_snopes_claims.py Snopes.xlsx --limit 20 -o /tmp/pilot.json

    # Full run:
    python scripts/enrich_snopes_claims.py Snopes.xlsx \\
        -o data/snopes/snopes_enriched.json

What this script does NOT do
-----------------------------
- It does not independently fact-check anything; it trusts Snopes' own
  normalised_rating for filtering, but every output claim is still
  tagged "verified": false pending this team's own review.
- It does not touch politifact_*/factcheck_* files — see
  scripts/README.md for their separate status.
- It is not part of the eiger package and is not covered by the 100%
  coverage gate (pyproject.toml's [tool.coverage.run] omits scripts/*).
"""

from __future__ import annotations

import argparse
import json
import sys
import time
from pathlib import Path
from typing import Any

try:
    import openpyxl
except ImportError as exc:
    raise SystemExit(
        "openpyxl is required for this script. Install it with: "
        "pip install 'eiger[data-import]' (or: pip install openpyxl)"
    ) from exc

from eiger.config import get_settings
from eiger.llm.ollama import DEFAULT_MODEL, OllamaLLM

_REQUIRED_COLUMNS = (
    "claim_id",
    "claim",
    "url",
    "date_published",
    "original_verdict",
    "normalised_rating",
)

_QUESTION_PROMPT_TEMPLATE = (
    "You will be given a single verified factual statement. Write ONE natural "
    "English question that this statement directly and fully answers. "
    "Reply with ONLY the question, no preamble, no quotation marks, no numbering.\n\n"
    "Statement: {claim}\n\n"
    "Question:"
)


def _read_raw_rows(path: Path) -> list[dict[str, Any]]:
    """Read every data row from the raw Snopes export, keyed by header name."""
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sheet = workbook[workbook.sheetnames[0]]

    rows_iter = sheet.iter_rows(values_only=True)
    header = next(rows_iter)
    header_index = {str(name): idx for idx, name in enumerate(header) if name is not None}

    missing = [col for col in _REQUIRED_COLUMNS if col not in header_index]
    if missing:
        raise SystemExit(
            f"'{path}' is missing expected column(s): {missing}. "
            f"Found: {list(header_index)}"
        )

    rows: list[dict[str, Any]] = []
    for raw_row in rows_iter:
        if raw_row is None:
            continue
        rows.append({name: raw_row[idx] for name, idx in header_index.items()})
    return rows


def filter_and_dedupe(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """
    Keep only ``normalised_rating is True`` rows, deduplicated by
    ``claim_id`` (first occurrence wins), preserving original file order.
    """
    seen_ids: set[Any] = set()
    kept: list[dict[str, Any]] = []
    for row in rows:
        if row.get("normalised_rating") is not True:
            continue
        claim_id = row.get("claim_id")
        if claim_id in seen_ids:
            continue
        seen_ids.add(claim_id)
        kept.append(row)
    return kept


def _load_existing(output_path: Path) -> tuple[list[dict[str, Any]], set[str]]:
    """Load already-enriched claims (if any) so a re-run can resume."""
    if not output_path.exists():
        return [], set()
    existing: list[dict[str, Any]] = json.loads(output_path.read_text(encoding="utf-8"))
    done_ids = {item["claim_id"] for item in existing}
    return existing, done_ids


def _to_claim_id(raw_claim_id: object) -> str:
    """
    Build the "SNOPES_<id>" claim_id, normalizing whole-number floats to
    int first.

    openpyxl reads numeric cells with no explicit integer format as Python
    floats (e.g. the raw claim_id column reads back as 80214.0, not
    80214), which would otherwise produce claim_ids like "SNOPES_80214.0"
    — cosmetically wrong and inconsistent with the EIB_CLAIM_NNN /
    EIB_CANDIDATE_NNN integer-suffix convention used elsewhere.
    """
    if isinstance(raw_claim_id, float) and raw_claim_id.is_integer():
        raw_claim_id = int(raw_claim_id)
    return f"SNOPES_{raw_claim_id}"


def generate_context_query(llm: OllamaLLM, claim_text: str) -> str:
    """Ask the LLM for one natural question the claim directly answers."""
    prompt = _QUESTION_PROMPT_TEMPLATE.format(claim=claim_text)
    question = llm.generate(prompt).strip()
    # Guard against a model wrapping the question in quotes despite instructions.
    return question.strip('"').strip("'").strip()


def _to_enriched_entry(row: dict[str, Any], llm: OllamaLLM) -> dict[str, Any]:
    context_query = generate_context_query(llm, str(row["claim"]))
    verdict = row["original_verdict"]
    published = row["date_published"]
    return {
        "claim_id": _to_claim_id(row["claim_id"]),
        "original_fact": str(row["claim"]),
        "context_query": context_query,
        "source": str(row["url"]),
        "notes": f"original_verdict={verdict}; date_published={published}",
        "verified": False,
    }


def enrich(
    input_path: Path,
    output_path: Path,
    *,
    model: str,
    host: str,
    port: int,
    limit: int | None,
    checkpoint_every: int,
) -> None:
    """Run the full filter -> dedupe -> enrich -> checkpoint pipeline."""
    rows = filter_and_dedupe(_read_raw_rows(input_path))
    if limit is not None:
        rows = rows[:limit]

    enriched, done_ids = _load_existing(output_path)
    remaining = [row for row in rows if _to_claim_id(row.get("claim_id")) not in done_ids]

    print(
        f"{len(rows)} candidate claim(s) after filtering; "
        f"{len(done_ids)} already enriched; {len(remaining)} remaining."
    )
    if not remaining:
        print("Nothing to do.")
        return

    llm = OllamaLLM(model_name=model, host=host, port=port, temperature=0.0)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    start = time.monotonic()
    for i, row in enumerate(remaining, start=1):
        enriched.append(_to_enriched_entry(row, llm))

        if i % checkpoint_every == 0 or i == len(remaining):
            output_path.write_text(
                json.dumps(enriched, indent=2, ensure_ascii=False) + "\n", encoding="utf-8"
            )
            elapsed = time.monotonic() - start
            print(
                f"  {i}/{len(remaining)} enriched ({elapsed:.0f}s elapsed) "
                f"-> checkpoint saved to {output_path}"
            )

    print(f"Done: {len(enriched)} total enriched claim(s) in {output_path}")


def main(argv: list[str] | None = None) -> int:
    """CLI entry point. Returns 0 on success, 1 if the input file is missing."""
    parser = argparse.ArgumentParser(
        description=(
            "Filter, deduplicate, and LLM-enrich a raw Snopes export into the "
            "JSON schema SnopesDataset expects. Requires a running Ollama server."
        )
    )
    parser.add_argument("input", help="Path to the raw Snopes .xlsx export.")
    parser.add_argument(
        "-o",
        "--output",
        default="data/snopes/snopes_enriched.json",
        help=(
            "Output path (default: data/snopes/snopes_enriched.json, "
            "matching SnopesDataset's default)."
        ),
    )
    parser.add_argument(
        "--limit", type=int, default=None, help="Cap the number of claims (pilot runs)."
    )
    parser.add_argument(
        "--checkpoint-every", type=int, default=25, help="Save progress every N claims."
    )
    parser.add_argument(
        "--model", default=DEFAULT_MODEL, help=f"Ollama model name (default: {DEFAULT_MODEL})."
    )
    parser.add_argument("--ollama-host", default=None, help="Override EIGER_OLLAMA_HOST.")
    parser.add_argument("--ollama-port", type=int, default=None, help="Override EIGER_OLLAMA_PORT.")
    args = parser.parse_args(argv)

    input_path = Path(args.input)
    if not input_path.exists():
        print(f"Error: file not found: {input_path}", file=sys.stderr)
        return 1

    settings = get_settings()
    enrich(
        input_path=input_path,
        output_path=Path(args.output),
        model=args.model,
        host=args.ollama_host or settings.ollama_host,
        port=args.ollama_port or settings.ollama_port,
        limit=args.limit,
        checkpoint_every=args.checkpoint_every,
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
