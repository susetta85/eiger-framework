"""
scripts/import_claims_xlsx.py

Convert a filled-in copy of ``eiger_claims_template.xlsx`` (the spreadsheet
handed to non-technical contributors — see the "Claim" sheet's "Istruzioni"
tab) into a JSON "candidate claims" file.

Why "candidate" and not eibench_raw_claims.json directly
----------------------------------------------------------
Claims collected via the spreadsheet are entered by a human but not yet
fact-checked against their cited source. This script performs ZERO
verification of its own — it only structures what was typed into the
sheet. Writing candidates to a separate file (never directly into
``eibench_raw_claims.json``, the fixture consumed by ``JSONFixtureDataset``
and exercised by the test suite / CI) guarantees that unverified content
can never silently leak into the framework's canonical regression fixture
or into a real experiment run.

The workflow this script is one step of:
  1. A contributor fills in eiger_claims_template.xlsx (no technical
     knowledge required).
  2. This script converts it to a candidate JSON file, tagging every
     claim "verified": false.
  3. A researcher checks each candidate's "source" field against the
     actual primary source and either discards it or manually copies the
     reviewed fields into eibench_raw_claims.json (mapping to the schema
     JSONFixtureDataset expects: claim_id/original_fact/context_query/
     adversarial_variants — see docs/DATASETS.md and
     eiger/datasets/README.md). There is no automatic promotion step;
     see scripts/README.md for the full rationale.

Expected spreadsheet structure (see eiger_claims_template.xlsx)
------------------------------------------------------------------
Sheet "Claim", columns A-E:
    A: Fatto verificato    (required)
    B: Domanda naturale    (required)
    C: Fonte               (recommended; not enforced by this script)
    D: Dominio             (free text — the template offers a dropdown,
                            but any string surviving a manual paste is
                            accepted here)
    E: Note (facoltativo)
Row 1 is the header. Row 2 is the template's own worked example and is
always skipped, regardless of its contents. Data starts at row 3.

Usage:
    python scripts/import_claims_xlsx.py claims.xlsx
    python scripts/import_claims_xlsx.py claims.xlsx -o data/candidates/batch1.json

Output JSON shape (one object per accepted row):
    {
      "claim_id": "EIB_CANDIDATE_001",
      "original_fact": "...",
      "context_query": "...",
      "source": "...",
      "domain": "...",
      "notes": "...",
      "verified": false
    }

This is intentionally NOT the schema JSONFixtureDataset expects — that
schema has no field for source/domain/verification status, and
conflating "just imported" with "ready to use in an experiment" is
exactly the mistake this staging format exists to prevent.

What this script does NOT do
-----------------------------
- It does not fact-check anything: "verified" is always False in its
  output; only a human sets it to True, in the promoted copy.
- It does not write to or modify eibench_raw_claims.json.
- It does not deduplicate against existing claims or previous candidate
  batches (each run's claim_ids restart at 001; keep batches in separate
  output files via -o if you run this more than once).
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any

try:
    import openpyxl
except ImportError as exc:
    raise SystemExit(
        "openpyxl is required for this script. Install it with: "
        "pip install 'eiger[data-import]' (or: pip install openpyxl)"
    ) from exc

_SHEET_NAME = "Claim"
_FIRST_DATA_ROW = 3  # row 1 = header, row 2 = the template's worked example
_CLAIM_ID_PREFIX = "EIB_CANDIDATE"

# 1-indexed column numbers, matching eiger_claims_template.xlsx exactly.
_COLUMNS = {
    "original_fact": 1,  # A: Fatto verificato
    "context_query": 2,  # B: Domanda naturale
    "source": 3,  # C: Fonte
    "domain": 4,  # D: Dominio
    "notes": 5,  # E: Note (facoltativo)
}


def _clean(value: object) -> str:
    """Coerce an Excel cell value to a stripped string, treating None as ''."""
    if value is None:
        return ""
    return str(value).strip()


def _read_rows(path: Path) -> list[dict[str, str]]:
    """
    Read every non-blank data row (from _FIRST_DATA_ROW onward) from the
    "Claim" sheet.

    Args:
        path: Path to the filled .xlsx file.

    Returns:
        List of raw {column_name: value} dicts, one per non-blank row, in
        file order. A row is considered blank (and skipped here) only if
        BOTH original_fact and context_query are empty; a row with just
        one of the two present is still returned, so the caller can flag
        it as invalid rather than silently dropping a half-filled entry.

    Raises:
        SystemExit: If the workbook has no sheet named "Claim".
    """
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if _SHEET_NAME not in workbook.sheetnames:
        raise SystemExit(
            f"Expected a sheet named '{_SHEET_NAME}' in '{path}', found: "
            f"{workbook.sheetnames}. Use eiger_claims_template.xlsx as-is and "
            "only fill in the 'Claim' sheet."
        )
    sheet = workbook[_SHEET_NAME]

    rows: list[dict[str, str]] = []
    for row in sheet.iter_rows(min_row=_FIRST_DATA_ROW):
        values = {name: _clean(row[col_idx - 1].value) for name, col_idx in _COLUMNS.items()}
        if not values["original_fact"] and not values["context_query"]:
            continue  # fully blank row — end of data or an accidental gap
        rows.append(values)
    return rows


def _to_candidate(row: dict[str, str], index: int) -> dict[str, Any]:
    """Build one candidate claim dict, generating a sequential claim_id."""
    return {
        "claim_id": f"{_CLAIM_ID_PREFIX}_{index:03d}",
        "original_fact": row["original_fact"],
        "context_query": row["context_query"],
        "source": row["source"],
        "domain": row["domain"],
        "notes": row["notes"],
        "verified": False,
    }


def convert(input_path: Path) -> tuple[list[dict[str, Any]], list[int]]:
    """
    Convert a filled xlsx template into candidate claim dicts.

    Args:
        input_path: Path to the filled .xlsx file.

    Returns:
        (candidates, invalid_row_numbers) — invalid_row_numbers are
        1-indexed spreadsheet row numbers (matching what the contributor
        sees in Excel) for rows missing original_fact or context_query,
        so they can be reported and fixed at the source rather than
        silently dropped or half-imported.
    """
    rows = _read_rows(input_path)

    candidates: list[dict[str, Any]] = []
    invalid_rows: list[int] = []
    next_index = 1
    for offset, row in enumerate(rows):
        sheet_row_number = _FIRST_DATA_ROW + offset
        if not row["original_fact"] or not row["context_query"]:
            invalid_rows.append(sheet_row_number)
            continue
        candidates.append(_to_candidate(row, next_index))
        next_index += 1

    return candidates, invalid_rows


def main(argv: list[str] | None = None) -> int:
    """CLI entry point. Returns 0 on success, 1 if the input file is missing."""
    parser = argparse.ArgumentParser(
        description=(
            "Convert a filled eiger_claims_template.xlsx into a JSON candidate "
            "claims file. Output claims are UNVERIFIED ('verified': false) and "
            "must be reviewed before being promoted into eibench_raw_claims.json."
        )
    )
    parser.add_argument("input", help="Path to the filled .xlsx file.")
    parser.add_argument(
        "-o",
        "--output",
        default="claims_candidate.json",
        help="Path to write the candidate JSON file (default: claims_candidate.json).",
    )
    args = parser.parse_args(argv)

    input_path = Path(args.input)
    if not input_path.exists():
        print(f"Error: file not found: {input_path}", file=sys.stderr)
        return 1

    candidates, invalid_rows = convert(input_path)

    output_path = Path(args.output)
    output_path.write_text(
        json.dumps(candidates, indent=2, ensure_ascii=False) + "\n", encoding="utf-8"
    )

    print(f"Imported {len(candidates)} candidate claim(s) -> {output_path}")
    if invalid_rows:
        print(
            f"Skipped {len(invalid_rows)} row(s) missing 'Fatto verificato' or "
            f"'Domanda naturale' (spreadsheet row numbers: {invalid_rows}). "
            "Fix these in the source file and re-run.",
            file=sys.stderr,
        )
    print(
        "All claims are UNVERIFIED. Review each one against its 'source' field "
        "before moving it into eibench_raw_claims.json."
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
